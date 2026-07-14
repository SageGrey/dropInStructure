import openpyxl
from collections import OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

F='5.2026_CompiledInvoices.xlsx'
wb=openpyxl.load_workbook(F); ws=wb['Sam_Daitzman']
G='Georgia'
TITLE=Font(name=G,bold=True,size=9); HDR=Font(name=G,bold=True,size=9,color='FFFFFF')
HDRFILL=PatternFill('solid',fgColor='434343'); TITLEFILL=PatternFill('solid',fgColor='FFE599')
NF=Font(name=G,size=9); BF=Font(name=G,bold=True,size=9)
TASKFILL=PatternFill('solid',fgColor='FCFCEA')
LIFILL=PatternFill('solid',fgColor='F3F3F3'); CATFILL=PatternFill('solid',fgColor='FFF2CC'); GTFILL=PatternFill('solid',fgColor='FFE599')
thin=Side(style='thin',color='BFBFBF'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
L=Alignment(horizontal='left',vertical='top',wrap_text=True); Cc=Alignment(horizontal='center',vertical='center'); Rc=Alignment(horizontal='right',vertical='center'); Lc=Alignment(horizontal='left',vertical='center')
HFMT='#,##0.00'; CUR='$#,##0.00'

# unmerge + clear from row 31 down
for m in list(ws.merged_cells.ranges):
    if m.min_row>=31: ws.unmerge_cells(str(m))
for r in range(31,90):
    for c in range(1,10):
        cell=ws.cell(r,c); cell.value=None; cell.fill=PatternFill(); cell.border=Border()
        cell.font=Font(name=G,size=9); cell.number_format='General'; cell.alignment=Alignment()

def title(r,t):
    ws.cell(r,1,t).font=TITLE
    for c in range(1,10): ws.cell(r,c).fill=TITLEFILL
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
def lblval(r,label,formula,fmt,vf=BF):
    ws.cell(r,1,label).font=BF; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    v=ws.cell(r,5,formula); v.font=vf; v.number_format=fmt; v.alignment=Rc

# groups cat->li->[unique tasks]
groups=OrderedDict()
for r in range(13,26):
    cat=ws.cell(r,3).value; li=ws.cell(r,4).value; task=ws.cell(r,5).value
    d=groups.setdefault(cat,OrderedDict()).setdefault(li,[])
    if task not in d: d.append(task)

title(31,'Task Breakdown')
for c,t in [(3,'Category'),(4,'Line Item'),(5,'Task'),(6,'Hours')]:
    cell=ws.cell(32,c,t); cell.font=HDR; cell.fill=HDRFILL; cell.border=bd; cell.alignment=Cc if c==6 else Lc

r=33; cat_subs=[]
for cat,lis in groups.items():
    li_subs=[]
    for li,tasks in lis.items():
        task_rows=[]
        for task in tasks:
            ws.cell(r,3,cat).font=NF; ws.cell(r,3).alignment=L
            ws.cell(r,4,li).font=NF; ws.cell(r,4).alignment=L
            ws.cell(r,5,task).font=NF; ws.cell(r,5).alignment=L
            ws.cell(r,6,f'=SUMIFS($I$13:$I$25,$C$13:$C$25,$C{r},$D$13:$D$25,$D{r},$E$13:$E$25,$E{r})').font=NF
            ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
            for c in range(3,7): ws.cell(r,c).border=bd; ws.cell(r,c).fill=TASKFILL
            task_rows.append(r); r+=1
        # line-item subtotal (always)
        ws.cell(r,4,f'{li} — line subtotal').font=BF; ws.cell(r,4).alignment=L
        ws.cell(r,6,f'=SUM(F{task_rows[0]}:F{task_rows[-1]})').font=BF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
        for c in range(3,7): ws.cell(r,c).fill=LIFILL; ws.cell(r,c).border=bd
        li_subs.append(r); r+=1
    ws.cell(r,3,f'{cat} — Category Subtotal').font=BF
    ws.cell(r,6,'='+'+'.join(f'F{x}' for x in li_subs)).font=BF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
    for c in range(3,7): ws.cell(r,c).fill=CATFILL; ws.cell(r,c).border=bd
    cat_subs.append(r); r+=1
ws.cell(r,3,'GRAND TOTAL').font=BF
ws.cell(r,6,'='+'+'.join(f'F{x}' for x in cat_subs)).font=BF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
for c in range(3,7): ws.cell(r,c).fill=GTFILL; ws.cell(r,c).border=bd
bk=r

# Software Costs
sr=bk+2; title(sr,'Software Costs')
ws.cell(sr+1,2,'Software').font=HDR; ws.cell(sr+1,2).fill=HDRFILL; ws.cell(sr+1,2).border=bd; ws.cell(sr+1,2).alignment=Lc
ws.cell(sr+1,5,'Cost').font=HDR; ws.cell(sr+1,5).fill=HDRFILL; ws.cell(sr+1,5).border=bd; ws.cell(sr+1,5).alignment=Cc
for rr in (sr+2,sr+3):
    for c in (2,5): ws.cell(rr,c).border=bd
    ws.cell(rr,5).number_format=CUR
ws.cell(sr+4,1,'Total Software Cost').font=BF; ws.merge_cells(start_row=sr+4,start_column=1,end_row=sr+4,end_column=3)
v=ws.cell(sr+4,5,f'=SUM(E{sr+2}:E{sr+3})'); v.font=BF; v.number_format=CUR; v.alignment=Rc
soft=sr+4
# Invoice Total
ir=soft+2; title(ir,'Invoice Total')
lblval(ir+1,'Total Labor Cost','=E29',CUR)
lblval(ir+2,'Total Software Cost',f'=E{soft}',CUR)
ws.cell(ir+3,1,'INVOICE TOTAL').font=Font(name=G,bold=True,size=10); ws.merge_cells(start_row=ir+3,start_column=1,end_row=ir+3,end_column=3)
v=ws.cell(ir+3,5,f'=E{ir+1}+E{ir+2}'); v.font=Font(name=G,bold=True,size=10); v.number_format=CUR; v.alignment=Rc
for c in range(1,6): ws.cell(ir+3,c).fill=GTFILL
last=ir+3
ws.print_area=f'A1:I{last}'
wb.save(F)
print('breakdown ends',bk,'| software',soft,'| invoice last',last)
