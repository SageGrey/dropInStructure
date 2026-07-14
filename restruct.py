import openpyxl
from copy import copy
from openpyxl.styles import PatternFill

F='5.2026_CompiledInvoices.xlsx'
wb=openpyxl.load_workbook(F)

# NRCS map by normalized contractor project name -> (pid, category, line_item)
def n(s): return ' '.join(str(s).strip().lower().split()) if s is not None else ''
PMAP={
 'aquacorridor upgrade':('AQU_2026_8','AQUATIC FRAMEWORK','Aqua Corridors Tool Suite'),
 'aquacorridors tool suite upgrades':('AQU_2026_8','AQUATIC FRAMEWORK','Aqua Corridors Tool Suite'),
 'southern waters':('AQU_2026_4','AQUATIC FRAMEWORK','Framework Development and Presentation'),
 'southern waters framework diagrams and presentation':('AQU_2026_4','AQUATIC FRAMEWORK','Framework Development and Presentation'),
 'helene workshop coordination':('AQU_2026_2','AQUATIC FRAMEWORK','Webinar and Workshop Support'),
 'helene workshop materials':('AQU_2026_3','AQUATIC FRAMEWORK',None),
 'aquatics photo blitz':('AQU_2026_5','AQUATIC FRAMEWORK','Framework Visuals and Posters'),
 'landowner willingness modeling':('AQU_2026_1','AQUATIC FRAMEWORK','Data Collection and Processing, For Coordinating Aquatic Framework'),
 'spring quarterly meeting with georgia team':('AQU_2026_6','AQUATIC FRAMEWORK','Framework Development and Presentation'),
 'anchor: conservation board upgrades':('LP_2026_2','LP-ANCHOR',None),
 'ple/plb outreah kit materials':('AQU_2026_7','AQUATIC FRAMEWORK','Framework Visuals and Posters'),
}
HEADERS=['date','client','category','line_item','task_fl_name','project_id_fl_name','notes','minutes','hours']
FLAG=PatternFill('solid',start_color='FFEB9C')

# per-sheet config: source col indices, data range, rate
CFG={
 'Jecha_Goss':  dict(hdr=11,d0=12,d1=19, c_date=1,c_client=2,c_proj=3,c_task=5,c_hours=6,c_pid=4, rate=80, fmt_hi=22),
 'Lucas_Goren': dict(hdr=11,d0=12,d1=25, c_date=1,c_client=2,c_proj=3,c_task=4,c_hours=5,c_pid=None, rate=80, fmt_hi=26),
 'Madeline_Cahue':dict(hdr=11,d0=12,d1=23, c_date=1,c_client=2,c_proj=3,c_task=4,c_hours=5,c_pid=None, rate=80, fmt_hi=23),
 'Elisa_Friedman':dict(hdr=11,d0=12,d1=22, c_date=1,c_client=2,c_proj=3,c_task=4,c_hours=5,c_pid=None, rate=80, fmt_hi=25),
 'Dieter_Brehm':dict(hdr=11,d0=12,d1=22, c_date=1,c_client=2,c_proj=3,c_task=4,c_hours=5,c_pid=None, rate=120, fmt_hi=22),
}

flagged={}
for name,cf in CFG.items():
    ws=wb[name]
    # 1. capture source rows
    src=[]
    for r in range(cf['d0'],cf['d1']+1):
        date=ws.cell(r,cf['c_date']).value
        client=ws.cell(r,cf['c_client']).value
        proj=ws.cell(r,cf['c_proj']).value
        task=ws.cell(r,cf['c_task']).value
        hours=ws.cell(r,cf['c_hours']).value
        pid_exist=ws.cell(r,cf['c_pid']).value if cf['c_pid'] else None
        src.append((r,date,client,proj,task,hours,pid_exist))
    # 2. header
    hf=copy(ws.cell(cf['hdr'],1).font); hfill=copy(ws.cell(cf['hdr'],1).fill); hal=copy(ws.cell(cf['hdr'],1).alignment)
    for c,h in enumerate(HEADERS,1):
        cell=ws.cell(cf['hdr'],c,h); cell.font=copy(hf); cell.fill=copy(hfill); cell.alignment=copy(hal)
        cell.number_format='General'
    # 3. clear body cols 1-9 across data..fmt_hi
    for r in range(cf['d0'],cf['fmt_hi']+1):
        for c in range(1,10): ws.cell(r,c).value=None
    # 4. write new rows
    flags=[]
    for (r,date,client,proj,task,hours,pid_exist) in src:
        key=n(proj)
        if key in PMAP:
            pid,cat,li=PMAP[key]
        elif pid_exist:
            pid,cat,li=str(pid_exist).strip(),None,None
        else:
            pid,cat,li=None,None,None
        projc=str(proj).strip() if proj else None
        clientc=str(client).strip() if client else None
        taskc=str(task).strip() if task else None
        mins=hours*60 if isinstance(hours,(int,float)) else None
        ws.cell(r,1).value=date
        ws.cell(r,2).value=clientc
        ws.cell(r,3).value=cat
        ws.cell(r,4).value=li
        ws.cell(r,5).value=projc
        ws.cell(r,6).value=pid
        ws.cell(r,7).value=taskc
        ws.cell(r,8).value=mins
        ws.cell(r,9).value=hours
        if pid is None:
            ws.cell(r,6).fill=copy(FLAG); flags.append((r,projc))
    flagged[name]=flags
    # 5. number formatting for the section
    for r in range(cf['d0'],cf['fmt_hi']+1):
        ws.cell(r,1).number_format='m/d/yyyy'
        for c in range(2,8): ws.cell(r,c).number_format='General'
        ws.cell(r,8).number_format='#,##0'
        ws.cell(r,9).number_format='#,##0.00'

wb.save(F)
print('data+headers+formatting done')
for k,v in flagged.items():
    print(k,'unmatched rows:',v)
