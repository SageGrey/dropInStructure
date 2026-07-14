import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F='5.2026_CompiledInvoices.xlsx'
wb=openpyxl.load_workbook(F)
ws=wb['Sam_Daitzman']

# ---------- style tokens ----------
G='Georgia'
TITLE=Font(name=G,bold=True,size=9)
HDR=Font(name=G,bold=True,size=9,color='FFFFFF')
HDRFILL=PatternFill('solid',fgColor='434343')
TITLEFILL=PatternFill('solid',fgColor='FFE599')
NF=Font(name=G,size=9)
BF=Font(name=G,bold=True,size=9)
INPUT=Font(name=G,size=9,color='0000FF')   # blue = editable input
BAND=PatternFill('solid',fgColor='F7F7F7')
LIFILL=PatternFill('solid',fgColor='F3F3F3')
CATFILL=PatternFill('solid',fgColor='FFF2CC')
GTFILL=PatternFill('solid',fgColor='FFE599')
thin=Side(style='thin',color='BFBFBF')
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
L=Alignment(horizontal='left',vertical='top',wrap_text=True)
Lc=Alignment(horizontal='left',vertical='center')
Rc=Alignment(horizontal='right',vertical='center')
Cc=Alignment(horizontal='center',vertical='center')
HFMT='#,##0.00'; CUR='$#,##0.00'

# ---------- 1. metadata tweak ----------
ws['A4']='Month: May 2026'

# ---------- 2. humanize labor header (row 12) ----------
labels=['Date','Client','Category','Line Item','Task','Project ID','Notes','Minutes','Hours']
for c,t in enumerate(labels,1):
    cell=ws.cell(12,c,t); cell.font=HDR; cell.fill=HDRFILL; cell.border=bd
    cell.alignment=Cc if c in (8,9) else Lc

# ---------- 3. style labor data rows 13-25 ----------
for i,r in enumerate(range(13,26)):
    band = BAND if i%2 else None
    for c in range(1,10):
        cell=ws.cell(r,c); cell.font=NF; cell.border=bd
        if band: cell.fill=band
        if c==1: cell.number_format='m/d/yyyy'; cell.alignment=Lc
        elif c==7: cell.alignment=L                       # notes wrap
        elif c==8: cell.number_format='#,##0'; cell.alignment=Cc
        elif c==9: cell.number_format=HFMT; cell.alignment=Cc
        else: cell.alignment=Alignment(horizontal='left',vertical='top',wrap_text=True)

# ---------- 4. unmerge + clear everything row 26+ ----------
for m in list(ws.merged_cells.ranges):
    if m.min_row>=26: ws.unmerge_cells(str(m))
for r in range(26,80):
    for c in range(1,10):
        cell=ws.cell(r,c)
        cell.value=None; cell.fill=PatternFill(); cell.border=Border()
        cell.font=Font(name=G,size=9); cell.number_format='General'
        cell.alignment=Alignment()

def title(r,text):
    ws.cell(r,1,text).font=TITLE
    for c in range(1,10): ws.cell(r,c).fill=TITLEFILL
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)

def lblval(r,label,formula,fmt,val_font=BF,lab_font=BF):
    ws.cell(r,1,label).font=lab_font
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
    v=ws.cell(r,5,formula); v.font=val_font; v.number_format=fmt; v.alignment=Rc

# ---------- 5. Labor summary (27-29) ----------
lblval(27,'Total Hours','=SUM(I13:I25)',HFMT)
lblval(28,'Labor Rate ($/hr)',160,CUR,val_font=INPUT)   # visible, editable
lblval(29,'Total Labor Cost','=E27*E28',CUR)

# ---------- 6. Task Breakdown (31-40) ----------
title(31,'Task Breakdown')
for c,t in [(3,'Category'),(4,'Line Item'),(5,'Task'),(6,'Hours')]:
    cell=ws.cell(32,c,t); cell.font=HDR; cell.fill=HDRFILL; cell.border=bd
    cell.alignment=Cc if c==6 else Lc
# rebuild groups from data
from collections import OrderedDict
groups=OrderedDict()
for r in range(13,26):
    cat=ws.cell(r,3).value; li=ws.cell(r,4).value; task=ws.cell(r,5).value
    groups.setdefault(cat,OrderedDict()).setdefault(li,[]).append(task)
    # dedupe tasks later
r=33; cat_sub=[]
for cat,lis in groups.items():
    cat_first=r
    for li,tasks in lis.items():
        seen=[]; 
        for t in tasks:
            if t not in seen: seen.append(t)
        li_first=r
        for task in seen:
            ws.cell(r,3,cat).font=NF; ws.cell(r,3).alignment=L
            ws.cell(r,4,li).font=NF; ws.cell(r,4).alignment=L
            ws.cell(r,5,task).font=NF; ws.cell(r,5).alignment=L
            f=f'=SUMIFS($I$13:$I$25,$C$13:$C$25,$C{r},$D$13:$D$25,$D{r},$E$13:$E$25,$E{r})'
            ws.cell(r,6,f).font=NF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
            for c in range(3,7): ws.cell(r,c).border=bd
            r+=1
        if len(seen)>1:
            ws.cell(r,4,f'{li} — subtotal').font=BF
            ws.cell(r,6,f'=SUM(F{li_first}:F{r-1})').font=BF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
            for c in range(3,7): ws.cell(r,c).fill=LIFILL; ws.cell(r,c).border=bd
            r+=1
    ws.cell(r,3,f'{cat} — Subtotal').font=BF
    ws.cell(r,6,f'=SUM(F{cat_first}:F{r-1})').font=BF; ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
    for c in range(3,7): ws.cell(r,c).fill=CATFILL; ws.cell(r,c).border=bd
    cat_sub.append(r); r+=1
ws.cell(r,3,'GRAND TOTAL').font=BF
ws.cell(r,6,'='+'+'.join(f'F{x}' for x in cat_sub)).font=BF
ws.cell(r,6).number_format=HFMT; ws.cell(r,6).alignment=Cc
for c in range(3,7): ws.cell(r,c).fill=GTFILL; ws.cell(r,c).border=bd
bk_end=r

# ---------- 7. Software Costs ----------
sr=bk_end+2
title(sr,'Software Costs')
ws.cell(sr+1,2,'Software').font=HDR; ws.cell(sr+1,2).fill=HDRFILL; ws.cell(sr+1,2).border=bd; ws.cell(sr+1,2).alignment=Lc
ws.cell(sr+1,5,'Cost').font=HDR; ws.cell(sr+1,5).fill=HDRFILL; ws.cell(sr+1,5).border=bd; ws.cell(sr+1,5).alignment=Cc
for rr in (sr+2,sr+3):           # two blank entry rows with borders
    for c in (2,5): ws.cell(rr,c).border=bd
    ws.cell(rr,5).number_format=CUR
ws.cell(sr+4,1,'Total Software Cost').font=BF
ws.merge_cells(start_row=sr+4,start_column=1,end_row=sr+4,end_column=3)
v=ws.cell(sr+4,5,f'=SUM(E{sr+2}:E{sr+3})'); v.font=BF; v.number_format=CUR; v.alignment=Rc
soft_total=sr+4

# ---------- 8. Invoice Total ----------
ir=soft_total+2
title(ir,'Invoice Total')
lblval(ir+1,'Total Labor Cost','=E29',CUR)
lblval(ir+2,'Total Software Cost',f'=E{soft_total}',CUR)
ws.cell(ir+3,1,'INVOICE TOTAL').font=Font(name=G,bold=True,size=10)
ws.merge_cells(start_row=ir+3,start_column=1,end_row=ir+3,end_column=3)
v=ws.cell(ir+3,5,f'=E{ir+1}+E{ir+2}'); v.font=Font(name=G,bold=True,size=10); v.number_format=CUR; v.alignment=Rc
for c in range(1,6): ws.cell(ir+3,c).fill=GTFILL
last=ir+3

# ---------- 9. column widths ----------
widths={'A':12,'B':10,'C':22,'D':30,'E':28,'F':14,'G':42,'H':9,'I':8}
for col,w in widths.items(): ws.column_dimensions[col].width=w

# ---------- 10. freeze panes + print area ----------
ws.freeze_panes='A13'
ws.print_area=f'A1:I{last}'
ws.page_setup.orientation='landscape'
ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr.fitToPage=True if ws.sheet_properties.pageSetUpPr else None

wb.save(F)
print('done. last row =',last,'| breakdown ends',bk_end,'| software total row',soft_total)
