import openpyxl
from collections import defaultdict, OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

F='5.2026_CompiledInvoices.xlsx'
wb=openpyxl.load_workbook(F)
ws=wb['Sam_Daitzman']

# 1. read labor data rows 13-41
rows=[]
for r in range(13,42):
    cat=ws.cell(r,3).value; li=ws.cell(r,4).value; task=ws.cell(r,5).value; hrs=ws.cell(r,9).value
    if task is None or not isinstance(hrs,(int,float)): continue
    rows.append((cat or '', li or '', task, hrs))

# 2. aggregate by (cat, li, task)
agg=OrderedDict()
for cat,li,task,h in rows:
    agg[(cat,li,task)]=agg.get((cat,li,task),0)+h

# 3. group cat -> li -> [tasks]
cats=OrderedDict()
for (cat,li,task) in agg:
    cats.setdefault(cat,OrderedDict()).setdefault(li,[]).append(task)

# styles
TITLE=Font(name='Georgia',bold=True,size=9)
HDR=Font(name='Georgia',bold=True,size=9,color='FFFFFF')
HDRFILL=PatternFill('solid',fgColor='434343')
NF=Font(name='Georgia',size=9)
SUBF=Font(name='Georgia',bold=True,size=9)
GTF=Font(name='Georgia',bold=True,size=9)
LIFILL=PatternFill('solid',fgColor='F3F3F3')
CATFILL=PatternFill('solid',fgColor='FFF2CC')
GTFILL=PatternFill('solid',fgColor='FFE599')
thin=Side(style='thin',color='BFBFBF')
bd=Border(left=thin,right=thin,top=thin,bottom=thin)
HFMT='#,##0.00'

# 4. clear old 45-56 (values + style)
for r in range(45,57):
    for c in range(1,9):
        cell=ws.cell(r,c); cell.value=None; cell.fill=PatternFill(); cell.border=Border()
        cell.font=Font(name='Georgia',size=9); cell.number_format='General'

# 5. write new section
r=45
ws.cell(r,1,'Task Breakdown').font=TITLE
r=46
hdr={3:'Category',4:'Line Item',5:'Task (task_fl_name)',6:'Hours'}
for c,t in hdr.items():
    cell=ws.cell(r,c,t); cell.font=HDR; cell.fill=HDRFILL; cell.border=bd
    cell.alignment=Alignment(horizontal='center' if c==6 else 'left')
cat_subtotal_rows=[]
r=47
for cat,lis in cats.items():
    cat_task_first=r
    for li,tasks in lis.items():
        li_first=r
        for task in tasks:
            ws.cell(r,3,cat).font=NF
            ws.cell(r,4,li).font=NF
            ws.cell(r,5,task).font=NF
            ws.cell(r,6,f'=SUMIFS($I$13:$I$41,$C$13:$C$41,$C{r},$D$13:$D$41,$D{r},$E$13:$E$41,$E{r})').font=NF
            ws.cell(r,6).number_format=HFMT
            for c in range(3,7): ws.cell(r,c).border=bd
            r+=1
        # line-item subtotal only if >1 task
        if len(tasks)>1:
            ws.cell(r,4,f'{li}  — subtotal').font=SUBF
            ws.cell(r,6,f'=SUM(F{li_first}:F{r-1})').font=SUBF; ws.cell(r,6).number_format=HFMT
            for c in range(3,7): ws.cell(r,c).fill=LIFILL; ws.cell(r,c).border=bd
            r+=1
    # category subtotal
    ws.cell(r,3,f'{cat} — Subtotal').font=SUBF
    ws.cell(r,6,f'=SUM(F{cat_task_first}:F{r-1})').font=SUBF; ws.cell(r,6).number_format=HFMT
    for c in range(3,7): ws.cell(r,c).fill=CATFILL; ws.cell(r,c).border=bd
    cat_subtotal_rows.append(r)
    r+=1
# grand total
ws.cell(r,3,'GRAND TOTAL').font=GTF
ws.cell(r,6,'='+'+'.join(f'F{x}' for x in cat_subtotal_rows)).font=GTF
ws.cell(r,6).number_format=HFMT
for c in range(3,7): ws.cell(r,c).fill=GTFILL; ws.cell(r,c).border=bd
last=r
wb.save(F)
print('Task Breakdown written, last row =',last)
print('categories:',list(cats.keys()))
